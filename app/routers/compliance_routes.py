# app/routers/compliance_routes.py
from fastapi import APIRouter, UploadFile, File, BackgroundTasks, HTTPException, Body, Request, Depends
from fastapi.responses import JSONResponse, StreamingResponse
from typing import Optional, Dict, Any, List
from bson import ObjectId
import traceback
from io import BytesIO
# Keep original relative imports (this file lives in app/routers/)
from ..compliance import run_full_pipeline, check_duplicate, aml_check_aadhaar
from ..security import get_current_user
from ..db import alerts_collection, audit_logs_collection, documents_collection, aml_blacklist_collection, kyc_data_collection, users_collection
from ..config import settings
import jwt

router = APIRouter(prefix="/compliance", tags=["compliance"])


# -----------------------
# NEW: List user documents
# -----------------------
@router.get("/docs", response_model=List[Dict[str, Any]])
def list_user_docs(current_user=Depends(get_current_user)):
    """
    Return documents uploaded by the current user ONLY.
    This applies to ALL users including admins.
    Admins can use the /submissions endpoint (Admin Panel) to see all users' docs.
    """
    try:
        user_id = None
        user_email = None
        try:
            if isinstance(current_user, dict):
                user_id = str(current_user.get("_id", ""))
                user_email = current_user.get("email", "")
            elif hasattr(current_user, "id"):
                user_id = str(current_user.id)
            elif hasattr(current_user, "_id"):
                user_id = str(current_user._id)
            elif isinstance(current_user, str):
                user_id = current_user
        except Exception:
            user_id = None

        # ALWAYS filter by current user - admins see their own docs in Submissions tab
        # They use Admin Panel for viewing all users' submissions
        if user_id:
            docs = list(documents_collection.find({"userId": user_id}).sort("createdAt", -1))
        elif user_email:
            docs = list(documents_collection.find({"userEmail": user_email}).sort("createdAt", -1))
        else:
            # No user context - return empty for security
            docs = []

        # normalize _id to string
        out = []
        for d in docs:
            dd = dict(d)
            dd["_id"] = str(dd.get("_id", dd.get("id", "")))
            out.append(dd)
        return out
    except Exception as e:
        tb = traceback.format_exc()
        print(tb)
        return JSONResponse(status_code=500, content={"error": str(e), "traceback": tb})


# -----------------------
# Existing endpoints (kept from original)
# -----------------------
@router.post("/verify_identity")
async def verify_identity(file: UploadFile = File(...), user_email: Optional[str] = None):
    """
    Run full KYC pipeline synchronously and return result.
    """
    try:
        content = await file.read()
        user = {"_id": user_email or "anonymous", "email": user_email}
        result = run_full_pipeline(user, file.filename, content)
        return result
    except Exception as e:
        tb = traceback.format_exc()
        print(tb)
        return JSONResponse(status_code=500, content={"error": str(e), "traceback": tb})


@router.post("/process_kyc")
async def process_kyc(background: BackgroundTasks, file: UploadFile = File(...), user_email: Optional[str] = None):
    """
    Start background KYC processing; returns immediately.
    """
    try:
        content = await file.read()
        user = {"_id": user_email or "anonymous", "email": user_email}
        background.add_task(run_full_pipeline, user, file.filename, content)
        return {"status": "processing"}
    except Exception as e:
        tb = traceback.format_exc()
        print(tb)
        return JSONResponse(status_code=500, content={"error": str(e), "traceback": tb})


@router.get("/fraud-score/{aadhaar}")
def fraud_score_for_aadhaar(aadhaar: str):
    try:
        docs = list(audit_logs_collection.find({"aadhaar": aadhaar}))
        scores = [d.get("fraud_score", 0) for d in docs if d.get("fraud_score") is not None]
        if scores:
            avg = sum(scores) / len(scores)
            label = "Low" if avg <= 30 else "Medium" if avg <= 70 else "High"
            return {"risk_score": int(avg), "risk_label": label, "source": "audit_logs"}
        return {"risk_score": 0, "risk_label": "Low", "source": "default"}
    except Exception as e:
        tb = traceback.format_exc()
        print(tb)
        return JSONResponse(status_code=500, content={"error": str(e), "traceback": tb})


@router.get("/aml/check/{aadhaar}")
def aml_check(aadhaar: str):
    try:
        res = aml_check_aadhaar(aadhaar)
        return res
    except Exception as e:
        tb = traceback.format_exc()
        print(tb)
        return JSONResponse(status_code=500, content={"error": str(e), "traceback": tb})


@router.get("/alerts")
def get_alerts():
    try:
        items = list(alerts_collection.find({"seen": {"$ne": True}}).sort("timestamp", -1))
        for i in items:
            i["_id"] = str(i["_id"])
        return items
    except Exception as e:
        tb = traceback.format_exc()
        print(tb)
        return JSONResponse(status_code=500, content={"error": str(e), "traceback": tb})


@router.post("/alerts/dismiss/{alert_id}")
def dismiss_alert_endpoint(alert_id: str):
    try:
        alerts_collection.update_one({"_id": ObjectId(alert_id)}, {"$set": {"seen": True}})
        return {"status": "dismissed"}
    except Exception as e:
        tb = traceback.format_exc()
        print(tb)
        return JSONResponse(status_code=500, content={"error": str(e), "traceback": tb})


@router.post("/logs/add")
def add_log(payload: Dict[str, Any] = Body(...)):
    try:
        res = audit_logs_collection.insert_one(payload)
        return {"ok": True, "id": str(res.inserted_id)}
    except Exception as e:
        tb = traceback.format_exc()
        print(tb)
        return JSONResponse(status_code=500, content={"error": str(e), "traceback": tb})


@router.get("/logs")
def get_logs():
    try:
        docs = list(audit_logs_collection.find().sort("createdAt", -1))
        for d in docs:
            d["_id"] = str(d["_id"])
        return docs
    except Exception as e:
        tb = traceback.format_exc()
        print(tb)
        return JSONResponse(status_code=500, content={"error": str(e), "traceback": tb})


@router.post("/check-duplicate")
def check_duplicate_endpoint(payload: Dict[str, Any] = Body(...)):
    try:
        aadhaar = payload.get("aadhaar")
        pan = payload.get("pan")
        res = check_duplicate(aadhaar, pan)
        return res
    except Exception as e:
        tb = traceback.format_exc()
        print(tb)
        return JSONResponse(status_code=500, content={"error": str(e), "traceback": tb})


@router.get("/documents/report")
def documents_report(request: Request, user_email: Optional[str] = None):
    """
    Generate a PDF report of uploaded documents. Optional `user_email` query param
    filters documents for a specific user.
    Returns: StreamingResponse with `application/pdf` and attachment header.
    """
    try:
        # If an Authorization Bearer token is provided, prefer its subject/email as the current user
        try:
            auth = request.headers.get("authorization") or request.headers.get("Authorization")
            if auth and auth.lower().startswith("bearer "):
                token = auth.split()[1]
                try:
                    payload = jwt.decode(token, settings.SECRET_KEY, algorithms=[settings.ALGORITHM])
                    token_sub = payload.get("sub") or payload.get("email")
                    if token_sub:
                        user_email = token_sub
                except Exception:
                    pass
        except Exception:
            pass

        # Build filter: if user_email provided, find document IDs for that user
        doc_id_set = None
        if user_email:
            docs_for_user = list(documents_collection.find({"userEmail": user_email}, {"_id": 1}))
            doc_id_set = set(str(d["_id"]) for d in docs_for_user)

        id_variants = None
        if doc_id_set is not None:
            id_variants = []
            from bson import ObjectId as _OID
            for sid in doc_id_set:
                id_variants.append(sid)
                try:
                    id_variants.append(_OID(sid))
                except Exception:
                    pass

        if id_variants is not None and len(id_variants) > 0:
            q = {"docId": {"$in": id_variants}}
        else:
            q = {}

        docs = list(kyc_data_collection.find(q).sort("createdAt", -1))

        source = 'kyc'
        if not docs:
            source = 'documents'
            if user_email:
                docs = list(documents_collection.find({"userEmail": user_email}).sort("createdAt", -1))
            else:
                docs = list(documents_collection.find().sort("createdAt", -1))

        try:
            from reportlab.lib.pagesizes import letter, landscape
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.units import inch
        except ImportError:
            return JSONResponse(status_code=500, content={"error": "reportlab is not installed. Run 'pip install reportlab'"})

        buf = BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(letter), leftMargin=40, rightMargin=40, topMargin=60, bottomMargin=40)
        styles = getSampleStyleSheet()
        small = ParagraphStyle("small", parent=styles["Normal"], fontSize=8)

        elements = []
        from datetime import datetime
        title = Paragraph("<b>Uploaded Documents Report</b>", styles["Title"])
        gen = Paragraph(f"Generated: {datetime.utcnow().isoformat()} UTC", styles["Normal"])
        user_header = Paragraph(f"User: <b>{user_email or 'N/A'}</b>", styles["Normal"])
        elements.extend([title, gen, user_header, Spacer(1, 12)])

        table_data = [["Doc ID", "Filename", "Type", "Created At", "Decision", "Fraud Score", "Fraud Reasons"]]

        def _format_created(c):
            try:
                if hasattr(c, "strftime"):
                    return c.strftime("%Y-%m-%d %H:%M")
                if isinstance(c, (int, float)):
                    from datetime import datetime as _dt
                    return _dt.utcfromtimestamp(c).strftime("%Y-%m-%d %H:%M")
                if isinstance(c, str):
                    s = c.strip()
                    if s.endswith("Z"):
                        s = s[:-1] + "+00:00"
                    try:
                        from datetime import datetime as _dt
                        dt = _dt.fromisoformat(s)
                        return dt.strftime("%Y-%m-%d %H:%M")
                    except Exception:
                        return s if len(s) <= 19 else s[:19] + "..."
            except Exception:
                pass
            return str(c)

        for d in docs:
            if source == 'kyc':
                doc_id = str(d.get("docId", ""))
                filename = d.get("verification", {}).get("filename") or ""
                if not filename:
                    try:
                        from bson import ObjectId as _OID
                        doc_obj = documents_collection.find_one({"_id": _OID(doc_id)})
                        if doc_obj:
                            filename = doc_obj.get("filename", "")
                    except Exception:
                        pass
                doc_type = d.get("docType", "")
                created = d.get("createdAt", "")
                decision = d.get("decision", "")
                fraud = d.get("fraud", {}) or {}
            else:
                doc_id = str(d.get("_id", ""))
                filename = d.get("filename", "")
                doc_type = d.get("docType", d.get("parsed", {}).get("aadhaarNumber") and "Aadhaar" or d.get("parsed", {}).get("panNumber") and "PAN" or "UNKNOWN")
                created = d.get("createdAt", "")
                decision = d.get("decision", d.get("verification", {}).get("decision", ""))
                fraud = d.get("fraud", {}) or d.get("verification", {}).get("fraud", {}) or {}

            fscore = fraud.get("score", "")
            freasons = fraud.get("reasons", [])
            if isinstance(freasons, list):
                freasons = ", ".join(str(x) for x in freasons)
            if not fscore and isinstance(fraud.get("details"), dict):
                fscore = fraud.get("details", {}).get("score", "")
                if not freasons:
                    freasons = fraud.get("details", {}).get("reasons", [])
                    if isinstance(freasons, list):
                        freasons = ", ".join(str(x) for x in freasons)

            doc_id_display = doc_id[:12] + "..." if len(doc_id) > 15 else doc_id
            filename = filename or ""
            doc_type = doc_type or ""
            created_display = _format_created(created or "")
            decision = decision or ""
            fscore = "" if fscore is None else str(fscore)
            freasons = freasons or ""

            row = [
                doc_id_display,
                Paragraph(filename, small),
                Paragraph(doc_type, small),
                Paragraph(created_display, small),
                Paragraph(decision, small),
                Paragraph(fscore, small),
                Paragraph(freasons, small),
            ]
            table_data.append(row)

        col_widths = [1.2 * inch, 2.5 * inch, 0.7 * inch, 1.2 * inch, 0.8 * inch, 0.6 * inch, 2.2 * inch]
        table = Table(table_data, colWidths=col_widths, repeatRows=1)
        table.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f2f4f8")),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ]))

        elements.append(table)
        doc.build(elements)
        buf.seek(0)
        headers = {"Content-Disposition": "attachment; filename=documents_report.pdf"}
        return StreamingResponse(buf, media_type="application/pdf", headers=headers)
    except Exception as e:
        tb = traceback.format_exc()
        print(tb)
        return JSONResponse(status_code=500, content={"error": str(e), "traceback": tb})


@router.get("/submissions")
def list_submissions(current_user=Depends(get_current_user)):
    """Return recent submissions (kyc snapshots preferred, fallback to documents)."""
    try:
        docs = list(kyc_data_collection.find().sort("createdAt", -1).limit(200))
        if not docs:
            docs = list(documents_collection.find().sort("createdAt", -1).limit(200))
            source = 'documents'
        else:
            source = 'kyc'

        # Cache user roles for efficiency
        user_roles_cache = {}
        
        def get_user_role(user_email):
            if not user_email:
                return 'user'
            if user_email in user_roles_cache:
                return user_roles_cache[user_email]
            user = users_collection.find_one({"email": user_email})
            role = user.get('role', 'user') if user else 'user'
            user_roles_cache[user_email] = role
            return role

        out = []
        for d in docs:
            user_email = d.get('userEmail')
            user_role = get_user_role(user_email)
            
            if source == 'kyc':
                doc_id = str(d.get('docId') or '')
                doc_obj = None
                try:
                    from bson import ObjectId as _OID
                    doc_obj = documents_collection.find_one({"_id": _OID(doc_id)})
                except Exception:
                    doc_obj = documents_collection.find_one({"_id": doc_id})
                record = {
                    'docId': doc_id,
                    'userEmail': user_email,
                    'userRole': user_role,
                    'filename': (doc_obj or {}).get('filename') if doc_obj else d.get('verification', {}).get('filename'),
                    'docType': d.get('docType'),
                    'createdAt': d.get('createdAt'),
                    'decision': d.get('decision'),
                    'fraud': d.get('fraud', {}),
                    'verification': d.get('verification', {}),
                }
            else:
                record = {
                    'docId': str(d.get('_id')),
                    'userEmail': user_email,
                    'userRole': user_role,
                    'filename': d.get('filename'),
                    'docType': d.get('docType'),
                    'createdAt': d.get('createdAt'),
                    'decision': d.get('decision'),
                    'fraud': d.get('fraud', {}),
                    'verification': d.get('verification', {}),
                }
            out.append(record)
        return out
    except Exception as e:
        tb = traceback.format_exc()
        print(tb)
        return JSONResponse(status_code=500, content={"error": str(e), "traceback": tb})


@router.post("/documents/{doc_id}/decision")
def set_document_decision(doc_id: str, payload: Dict[str, Any] = Body(...), current_user=Depends(get_current_user)):
    """Admin endpoint: set decision for a document (Approve/Reject).
    payload: { decision: 'Approve'|'Reject', notes: optional }
    """
    try:
        decision = payload.get('decision')
        notes = payload.get('notes', '')
        if decision not in ("Approve", "Reject"):
            return JSONResponse(status_code=400, content={"error": "invalid decision"})

        updated = documents_collection.update_one({"_id": ObjectId(doc_id)}, {"$set": {"decision": decision, "reviewer": current_user.get('email'), "reviewedAt": __import__('datetime').datetime.utcnow()}})
        kyc_data_collection.update_many({"docId": {"$in": [doc_id, ObjectId(doc_id)]}}, {"$set": {"decision": decision, "reviewer": current_user.get('email'), "reviewedAt": __import__('datetime').datetime.utcnow()}})

        audit_logs_collection.insert_one({
            "action": "document_decision",
            "docId": doc_id,
            "decision": decision,
            "notes": notes,
            "userEmail": current_user.get('email'),
            "createdAt": __import__('datetime').datetime.utcnow()
        })

        return {"ok": True, "updated": updated.modified_count}
    except Exception as e:
        tb = traceback.format_exc()
        print(tb)
        return JSONResponse(status_code=500, content={"error": str(e), "traceback": tb})


# -----------------------
# NEW: Bulk Verification from Excel/CSV
# -----------------------
@router.post("/bulk-verify")
async def bulk_verify_excel(file: UploadFile = File(...), current_user=Depends(get_current_user)):
    """
    Accept an Excel (.xlsx) or CSV file with KYC data and validate each row.
    Expected columns: Name, Aadhaar, PAN, DOB, Address
    Returns validation results for each row.
    """
    try:
        import re
        from datetime import datetime
        
        content = await file.read()
        filename = file.filename.lower()
        
        rows = []
        
        # Parse CSV
        if filename.endswith('.csv'):
            import csv
            from io import StringIO
            text = content.decode('utf-8', errors='ignore')
            reader = csv.DictReader(StringIO(text))
            for row in reader:
                rows.append(dict(row))
        
        # Parse Excel
        elif filename.endswith('.xlsx') or filename.endswith('.xls'):
            try:
                import openpyxl
                from io import BytesIO as BIO
                wb = openpyxl.load_workbook(BIO(content), read_only=True)
                ws = wb.active
                headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    row_dict = dict(zip(headers, row))
                    rows.append(row_dict)
            except ImportError:
                return JSONResponse(status_code=400, content={"error": "openpyxl not installed. Use CSV format or install openpyxl."})
        else:
            return JSONResponse(status_code=400, content={"error": "Unsupported file format. Use .csv or .xlsx"})
        
        if not rows:
            return JSONResponse(status_code=400, content={"error": "No data rows found in file"})
        
        results = []
        
        for idx, row in enumerate(rows):
            # Normalize column names (case-insensitive)
            norm_row = {k.lower().strip() if k else '': v for k, v in row.items()}
            
            name = norm_row.get('name', '')
            aadhaar = str(norm_row.get('aadhaar', '') or '').replace(' ', '').replace('-', '')
            pan = str(norm_row.get('pan', '') or '').upper().replace(' ', '')
            dob = str(norm_row.get('dob', '') or '')
            address = str(norm_row.get('address', '') or '')
            
            errors = []
            warnings = []
            fraud_score = 0
            
            # Validate Aadhaar (12 digits)
            if aadhaar:
                if not re.match(r'^\d{12}$', aadhaar):
                    errors.append("Invalid Aadhaar format (must be 12 digits)")
                    fraud_score += 30
                else:
                    # Check for duplicate
                    if documents_collection.find_one({"parsed.aadhaarNumber": aadhaar}):
                        warnings.append("Aadhaar already exists in system")
                        fraud_score += 20
            else:
                warnings.append("Aadhaar not provided")
            
            # Validate PAN (5 letters + 4 digits + 1 letter)
            if pan:
                if not re.match(r'^[A-Z]{5}\d{4}[A-Z]$', pan):
                    errors.append("Invalid PAN format (must be ABCDE1234F)")
                    fraud_score += 25
                else:
                    if documents_collection.find_one({"parsed.panNumber": pan}):
                        warnings.append("PAN already exists in system")
                        fraud_score += 20
            else:
                warnings.append("PAN not provided")
            
            # Validate DOB
            if dob:
                try:
                    # Try common formats
                    for fmt in ['%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d', '%d/%m/%y']:
                        try:
                            parsed_date = datetime.strptime(dob.strip(), fmt)
                            age = (datetime.now() - parsed_date).days // 365
                            if age < 18:
                                errors.append(f"Underage: {age} years old")
                                fraud_score += 40
                            break
                        except ValueError:
                            continue
                except Exception:
                    warnings.append("Could not parse DOB")
            
            # Validate Name
            if not name or len(name.strip()) < 3:
                errors.append("Name is required (min 3 characters)")
                fraud_score += 10
            
            # Determine status
            if errors:
                status = "Failed"
            elif warnings:
                status = "Review"
            else:
                status = "Pass"
            
            results.append({
                "row": idx + 1,
                "filename": f"Row {idx + 1}",
                "name": name,
                "aadhaar": f"XXXX-XXXX-{aadhaar[-4:]}" if len(aadhaar) >= 4 else aadhaar,
                "pan": f"XXXXX{pan[-5:]}" if len(pan) >= 5 else pan,
                "success": status != "Failed",
                "status": status,
                "fraudScore": min(fraud_score, 100),
                "decision": status,
                "errors": errors,
                "warnings": warnings,
            })
        
        # Summary
        passed = len([r for r in results if r['status'] == 'Pass'])
        review = len([r for r in results if r['status'] == 'Review'])
        failed = len([r for r in results if r['status'] == 'Failed'])
        
        return {
            "message": f"Processed {len(results)} rows",
            "summary": {"passed": passed, "review": review, "failed": failed, "total": len(results)},
            "results": results
        }
    except Exception as e:
        tb = traceback.format_exc()
        print(tb)
        return JSONResponse(status_code=500, content={"error": str(e), "traceback": tb})
